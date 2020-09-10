import openpyxl, os
from datetime import datetime

#change current working directory to database folder
os.chdir('C:\\Users\\Kwesi Joe\\Documents\\VScode\\face recognition')

#load workboook
database = openpyxl.load_workbook('Database.xlsx')

#open sheet
sheet = database['Sheet1']

#get last row of the number plate column
last_row = len(sheet['A'])

#if the car in going into the car park, append registration number
sheet['A'+str(last_row+1)]= 'licplate'

#record arrival time
sheet['B'+str(last_row+1)] = str(datetime.now())

#if the car is going out, check row where its number got recorded and record the departure time
for row in range(1,len(sheet['A'])):
    if sheet.cell(row,1).value == 'licplate' and sheet.cell(row,3) == None:
        sheet['C'+str(row)] = str(datetime.now())


#save the database

database.save('Database.xlsx')

